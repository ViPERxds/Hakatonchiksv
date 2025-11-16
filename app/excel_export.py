from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from datetime import datetime
from typing import List
from app.database import Invoice
import io


class ExcelExporter:
    @staticmethod
    def export_invoices(db: Session, user_id: int, invoices: List[Invoice] = None) -> io.BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Счета"
        
        headers = [
            'ID', 'Номер счета', 'Дата', 'Продавец', 'Покупатель',
            'ИНН', 'КПП', 'Сумма', 'Валюта', 'НДС', 'Дата обработки'
        ]
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        if invoices is None:
            invoices = db.query(Invoice).filter(
                Invoice.user_id == user_id
            ).order_by(Invoice.created_at.desc()).all()
        
        for row_num, invoice in enumerate(invoices, 2):
            data = invoice.extracted_data or {}
            vat_data = data.get('vat_amount', {})
            
            row_data = [
                invoice.id,
                invoice.invoice_number or data.get('invoice_number', ''),
                invoice.date or data.get('date', ''),
                invoice.seller or data.get('seller', ''),
                data.get('buyer', ''),
                data.get('seller_inn', ''),
                data.get('seller_kpp', ''),
                invoice.total_amount or data.get('total_amount', ''),
                invoice.currency or data.get('currency', 'RUB'),
                f"{vat_data.get('rate', '')}% - {vat_data.get('amount', '')}" if isinstance(vat_data, dict) else '',
                invoice.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def export_statistics(db: Session, user_id: int, stats: dict) -> io.BytesIO:
        wb = Workbook()
        
        ws1 = wb.active
        ws1.title = "Общая статистика"
        
        ws1['A1'] = 'Показатель'
        ws1['B1'] = 'Значение'
        ws1['A1'].font = Font(bold=True)
        ws1['B1'].font = Font(bold=True)
        
        ws1['A2'] = 'Всего счетов'
        ws1['B2'] = stats.get('total_invoices', 0)
        ws1['A3'] = 'Общая сумма'
        ws1['B3'] = stats.get('total_amount', 0)
        ws1['A4'] = 'Период (дней)'
        ws1['B4'] = stats.get('period_days', 30)
        
        ws2 = wb.create_sheet("Топ поставщиков")
        ws2['A1'] = 'Поставщик'
        ws2['B1'] = 'Количество счетов'
        ws2['A1'].font = Font(bold=True)
        ws2['B1'].font = Font(bold=True)
        
        for idx, seller in enumerate(stats.get('top_sellers', []), 2):
            ws2[f'A{idx}'] = seller.get('name', '')
            ws2[f'B{idx}'] = seller.get('count', 0)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

